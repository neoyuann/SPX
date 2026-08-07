"""One-time (or whenever-you-update-the-sheet) bulk import: reads a
Bloomberg-style roster spreadsheet — one sheet per market, a bold row
inside each sheet marking the start of a sub-sector, then Company Name /
Ticker rows below it — and adds every company it finds to companies.yaml.

Companies already tracked (matched by the ticker this script would derive)
are left with their existing sources untouched; only their `country` and
`sub_sector` fields are filled in or corrected from the sheet, so a
previously hand-tuned entry (real RSS feed, CIK, custom selectors) never
gets clobbered by a bulk run.

    python import_excel_companies.py "Company_Name_Tickers_and_Sub_Categories.xlsx"

Requires openpyxl (`pip install openpyxl`) — only this importer needs it,
the running tracker doesn't.
"""
from __future__ import annotations

import re
import sys

import openpyxl
import yaml

from tracker.addco import _find_block, _read, _validate, _write_atomic, yaml_scalar

COMPANIES_PATH = "companies.yaml"

# sheet name -> (region code, generic exchange label)
MARKET_INFO = {
    "Korea": ("KR", "KRX"),
    "Japan": ("JP", "TSE"),
    "China": ("CN", "SSE / SZSE"),
    "Taiwan": ("TW", "TWSE"),
    "Hong Kong": ("HK", "HKEX"),
    "US": ("US", "NYSE / NASDAQ"),
}

_TICKER_RE = re.compile(r"^(.+?)\s+([A-Z0-9]+)\s+Equity$")


def _derive_ticker(raw_ticker: str) -> str | None:
    m = _TICKER_RE.match(raw_ticker.strip())
    if not m:
        return None
    code, suffix = m.group(1).strip(), m.group(2).strip()
    if suffix == "US":
        return code.upper()
    if suffix == "HK":
        return f"{code.zfill(4)}.HK"
    if suffix == "TT":
        return f"{(code.zfill(4) if code.isdigit() else code)}.TW"
    if suffix == "JP":
        return f"{code}.T"
    if suffix == "KS":
        return f"{code}.KS"
    if suffix == "CH":
        return f"{code}.CH"
    return None  # unknown suffix — skip rather than guess


def parse_workbook(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    rows = []
    seen_tickers = set()
    for sheet_name in wb.sheetnames:
        if sheet_name not in MARKET_INFO:
            print(f"  skipping unrecognised sheet {sheet_name!r}")
            continue
        ws = wb[sheet_name]
        current_sub = "Unclassified"
        for row in ws.iter_rows(min_row=2, values_only=True):
            name, raw_ticker = (row + (None, None))[:2]
            if name is None and raw_ticker is not None:
                m = re.match(r"^(.*?)\s*\(\d+\)\s*$", str(raw_ticker).strip())
                current_sub = m.group(1).strip() if m else str(raw_ticker).strip()
                continue
            if name is None or raw_ticker is None:
                continue
            ticker = _derive_ticker(str(raw_ticker))
            if ticker is None:
                print(f"  couldn't parse ticker {raw_ticker!r} for {name!r}, skipping")
                continue
            if ticker in seen_tickers:
                continue  # e.g. a company listed under two sheets — keep the first
            seen_tickers.add(ticker)
            rows.append({
                # Kept as the sheet has it (Bloomberg-style short names, all
                # caps and sometimes truncated) rather than guessed at with
                # .title() — safer than mangling names like "KB FINANCIAL GRO".
                "ticker": ticker, "name": str(name).strip(),
                "country": sheet_name, "sub_sector": current_sub,
            })
    return rows


def _patch_existing(text: str, ticker: str, country: str, sub_sector: str) -> str:
    """Fill in / correct country and sub_sector on an already-tracked
    company without touching anything else in its block."""
    m = _find_block(text, ticker)
    block = m.group(1)
    country_q, sub_sector_q = yaml_scalar(country), yaml_scalar(sub_sector)
    if re.search(r"\n[ \t]*country:\s*.*", block):
        block = re.sub(r"(\n[ \t]*country:\s*).*", lambda _m: _m.group(1) + country_q, block, count=1)
    else:
        block += f"\n    country: {country_q}"
    if re.search(r"\n[ \t]*sub_sector:\s*.*", block):
        block = re.sub(r"(\n[ \t]*sub_sector:\s*).*", lambda _m: _m.group(1) + sub_sector_q, block, count=1)
    else:
        block += f"\n    sub_sector: {sub_sector_q}"
    return text[:m.start(1)] + block + text[m.end(1):]


def _new_block(row: dict) -> str:
    ticker, name, country, sub_sector = row["ticker"], row["name"], row["country"], row["sub_sector"]
    region, exchange = MARKET_INFO[country]
    lines = [
        f"\n  - ticker: {yaml_scalar(ticker)}",
        f"    name: {yaml_scalar(name)}",
        f"    exchange: {yaml_scalar(exchange)}",
        f"    region: {yaml_scalar(region)}",
        f"    country: {yaml_scalar(country)}",
        f"    sub_sector: {yaml_scalar(sub_sector)}",
        "    enabled: true",
        "    aliases: []",
        "    sources:",
    ]
    if country == "US":
        # No CIK on file — resolved automatically from the ticker at fetch
        # time against SEC's published ticker list (see tracker/sources.py).
        lines += ["      - type: sec_edgar", "        tier: regulatory"]
    lines += ["      - type: news", "        tier: news"]
    return "\n".join(lines) + "\n"


def main():
    if len(sys.argv) != 2:
        print(f"usage: python {sys.argv[0]} <path-to-xlsx>")
        sys.exit(1)
    xlsx_path = sys.argv[1]

    print(f"Reading {xlsx_path} ...")
    rows = parse_workbook(xlsx_path)
    print(f"Parsed {len(rows)} companies across {len(MARKET_INFO)} markets.")

    text = _read(COMPANIES_PATH)
    added = updated = 0
    for row in rows:
        if _find_block(text, row["ticker"]):
            text = _patch_existing(text, row["ticker"], row["country"], row["sub_sector"])
            updated += 1
        else:
            text = text.rstrip("\n") + "\n" + _new_block(row)
            added += 1

    ok, msg = _validate(text)
    if not ok:
        print(f"Import produced invalid YAML, nothing was written: {msg}")
        sys.exit(1)

    _write_atomic(COMPANIES_PATH, text)
    print(f"Done: {added} companies added, {updated} existing companies had "
          f"country/sub_sector filled in. {COMPANIES_PATH} now tracks "
          f"{added + updated} companies from this sheet (plus any tracked outside it).")


if __name__ == "__main__":
    main()
